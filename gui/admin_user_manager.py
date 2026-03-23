"""
用户管理模块
重写后的用户管理界面，采用现代化玻璃拟态设计风格
"""
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, 
    QLabel, QTableWidget, QTableWidgetItem, QHeaderView, 
    QMessageBox, QLineEdit, QDialog, QComboBox, QFrame,
    QCheckBox, QGraphicsDropShadowEffect, QAbstractItemView, 
    QDateTimeEdit, QSpinBox, QScrollArea, QSizePolicy,
    QStyledItemDelegate, QStyleOptionViewItem
)
from PyQt6.QtCore import Qt, pyqtSignal, QSize, QDateTime, QPropertyAnimation, QEasingCurve, QTimer, QRect
from PyQt6.QtGui import QFont, QColor, QIcon, QPainter, QLinearGradient, QPen, QBrush, QPainterPath
from database import DatabaseManager, User, Device, Card
from gui.styles import COLORS
from gui.icons import Icons
import datetime


# 扩展颜色系统 - 更丰富的渐变色
PREMIUM_COLORS = {
    **COLORS,
    # 渐变色组
    'gradient_blue_start': '#667eea',
    'gradient_blue_end': '#764ba2',
    'gradient_green_start': '#11998e',
    'gradient_green_end': '#38ef7d',
    'gradient_orange_start': '#f093fb',
    'gradient_orange_end': '#f5576c',
    'gradient_purple_start': '#4facfe',
    'gradient_purple_end': '#00f2fe',
    'gradient_gold_start': '#f7971e',
    'gradient_gold_end': '#ffd200',
    
    # 玻璃效果
    'glass_bg': 'rgba(255, 255, 255, 0.85)',
    'glass_border': 'rgba(255, 255, 255, 0.6)',
    'glass_shadow': 'rgba(31, 38, 135, 0.07)',
    
    # 深色点缀
    'dark_accent': '#1a1a2e',
    'text_heading': '#2d3748',
    'text_body': '#4a5568',
    'text_hint': '#a0aec0',
    
    # 功能色
    'mint': '#00d9a6',
    'coral': '#ff6b6b',
    'lavender': '#a29bfe',
    'sky': '#74b9ff',
}


class GlassFrame(QFrame):
    """玻璃拟态框架"""
    
    def __init__(self, parent=None, opacity=0.9, radius=24, hover_effect=False):
        super().__init__(parent)
        self.opacity = opacity
        self.radius = radius
        self.hover_effect = hover_effect
        self._setup_style()
    
    def _setup_style(self):
        self.setStyleSheet(f"""
            GlassFrame {{
                background: rgba(255, 255, 255, {self.opacity});
                border: 1px solid rgba(255, 255, 255, 0.8);
                border-radius: {self.radius}px;
            }}
            GlassFrame:hover {{
                background: rgba(255, 255, 255, {min(1.0, self.opacity + 0.05)});
                border-color: rgba(255, 255, 255, 1.0);
            }}
        """ if self.hover_effect else f"""
            GlassFrame {{
                background: rgba(255, 255, 255, {self.opacity});
                border: 1px solid rgba(255, 255, 255, 0.6);
                border-radius: {self.radius}px;
            }}
        """)
        
        # 添加高级阴影效果
        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(30)
        shadow.setColor(QColor(31, 38, 135, 15))
        shadow.setOffset(0, 8)
        self.setGraphicsEffect(shadow)


class GradientButton(QPushButton):
    """渐变按钮"""
    
    def __init__(self, text, start_color, end_color, parent=None):
        super().__init__(text, parent)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setFixedHeight(44)
        self.setStyleSheet(f"""
            QPushButton {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {start_color}, stop:1 {end_color});
                color: white;
                border: none;
                border-radius: 22px;
                font-weight: 600;
                font-size: 14px;
                padding: 0 24px;
            }}
            QPushButton:hover {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {end_color}, stop:1 {start_color});
            }}
            QPushButton:pressed {{
                padding-top: 2px;
            }}
        """)


class UserEditDialog(QDialog):
    """用户编辑/添加对话框 - 重新设计版"""
    
    def __init__(self, parent=None, user=None):
        super().__init__(parent)
        self.user = user
        self.init_ui()
        
    def init_ui(self):
        mode = "编辑用户" if self.user else "添加用户"
        self.setWindowTitle(mode)
        self.setFixedSize(460, 620)
        
        # 设置白色背景
        self.setStyleSheet("QDialog { background-color: white; }")
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        
        # === 1. 顶部 Header ===
        header = QFrame()
        header.setFixedHeight(110)
        header.setStyleSheet(f"""
            QFrame {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 {PREMIUM_COLORS['gradient_blue_start']}, 
                    stop:1 {PREMIUM_COLORS['gradient_blue_end']});
            }}
        """)
        
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(32, 0, 32, 0)
        header_layout.setSpacing(20)
        
        # 标题文字
        title_info = QVBoxLayout()
        title_info.setSpacing(6)
        title_info.setAlignment(Qt.AlignmentFlag.AlignVCenter)
        
        title_lbl = QLabel(mode)
        title_lbl.setStyleSheet("color: white; font-size: 26px; font-weight: 800;")
        
        subtitle_lbl = QLabel("请填写下方的用户信息表单" if not self.user else f"正在编辑 {self.user.username} 的信息")
        subtitle_lbl.setStyleSheet("color: rgba(255,255,255,0.85); font-size: 13px; font-weight: 500;")
        
        title_info.addWidget(title_lbl)
        title_info.addWidget(subtitle_lbl)
        
        # 图标
        icon_bg = QLabel("✏️" if self.user else "➕")
        icon_bg.setFixedSize(56, 56)
        icon_bg.setAlignment(Qt.AlignmentFlag.AlignCenter)
        icon_bg.setStyleSheet("""
            background: rgba(255,255,255,0.2);
            border-radius: 28px;
            font-size: 26px;
        """)
        
        header_layout.addLayout(title_info)
        header_layout.addWidget(icon_bg)
        
        layout.addWidget(header)
        
        # === 2. 表单区域 ===
        form_widget = QWidget()
        form_layout = QVBoxLayout(form_widget)
        form_layout.setContentsMargins(32, 32, 32, 20)
        form_layout.setSpacing(20)
        
        # 定义通用样式
        self.input_style = f"""
            QLineEdit, QComboBox, QDateTimeEdit, QSpinBox {{
                border: 1px solid {PREMIUM_COLORS['border_light']};
                border-radius: 10px;
                padding: 0 12px;
                background: #f8fafc;
                height: 42px;
                font-size: 14px;
                color: {PREMIUM_COLORS['text_heading']};
                selection-background-color: {PREMIUM_COLORS['primary_light']};
            }}
            QLineEdit:focus, QComboBox:focus, QDateTimeEdit:focus, QSpinBox:focus {{
                border: 1px solid {PREMIUM_COLORS['gradient_blue_start']};
                background: white;
            }}
            QLineEdit:hover, QComboBox:hover, QDateTimeEdit:hover, QSpinBox:hover {{
                background: white;
                border-color: #cbd5e1;
            }}
            QLineEdit:disabled, QComboBox:disabled, QDateTimeEdit:disabled, QSpinBox:disabled {{
                background: #e2e8f0;
                color: #94a3b8;
                border-color: #e2e8f0;
            }}
            QComboBox::drop-down {{ border: none; width: 24px; }}
            QDateTimeEdit::drop-down {{
                border: none;
                width: 30px;
                subcontrol-position: right center;
            }}
            QDateTimeEdit::down-arrow {{
                image: none;
                width: 16px;
                height: 16px;
            }}
        """
        form_widget.setStyleSheet(self.input_style)
        
        # 用户名
        self.username_input = QLineEdit()
        self.username_input.setPlaceholderText("请输入用户名")
        if self.user:
            self.username_input.setText(self.user.username)
            self.username_input.setEnabled(False)
            self.username_input.setStyleSheet(self.username_input.styleSheet() + "QLineEdit { background: #f1f5f9; color: #94a3b8; }")
        form_layout.addLayout(self._create_field("用户名", self.username_input, "用于登录系统的唯一标识"))
        
        # 密码 & 角色 (一行两列)
        row1 = QHBoxLayout()
        row1.setSpacing(20)
        
        self.password_input = QLineEdit()
        self.password_input.setPlaceholderText("留空不修改" if self.user else "设置密码")
        self.password_input.setEchoMode(QLineEdit.EchoMode.Password)
        row1.addLayout(self._create_field("密码", self.password_input, "用户登录凭证"))
        
        self.role_combo = QComboBox()
        self.role_combo.blockSignals(True)
        self.role_combo.model().blockSignals(True)
        self.role_combo.addItems(['user', 'admin'])
        self.role_combo.model().blockSignals(False)
        self.role_combo.blockSignals(False)
        if self.user: self.role_combo.setCurrentText(self.user.role)
        row1.addLayout(self._create_field("角色", self.role_combo, "admin拥有所有权限"))
        
        form_layout.addLayout(row1)
        
        # 有效期 & 使用次数 (一行两列)
        row2 = QHBoxLayout()
        row2.setSpacing(20)
        
        # 左：有效期 (使用容器控制宽度)
        expire_container = QWidget()
        expire_box = QVBoxLayout(expire_container)
        expire_box.setContentsMargins(0, 0, 0, 0)
        expire_box.setSpacing(8)
        expire_header = QHBoxLayout()
        expire_label = QLabel("账户有效期")
        expire_label.setStyleSheet(f"color: {PREMIUM_COLORS['text_body']}; font-weight: 600; font-size: 13px;")
        self.expire_check = QCheckBox()
        self.expire_check.setStyleSheet(self._get_check_style())
        expire_header.addWidget(expire_label)
        expire_header.addStretch()
        expire_header.addWidget(self.expire_check)
        
        self.expire_edit = QDateTimeEdit()
        self.expire_edit.setCalendarPopup(True)
        self.expire_edit.setDisplayFormat("yyyy-MM-dd")
        self.expire_edit.setDateTime(QDateTime.currentDateTime().addDays(5))
        self.expire_edit.setEnabled(False)
        
        # 设置日历弹出窗口样式
        self._setup_calendar_style(self.expire_edit)
        
        self.expire_check.stateChanged.connect(
            lambda state: self.expire_edit.setEnabled(state == Qt.CheckState.Checked.value)
        )
        if self.user and self.user.expire_time:
            # 编辑用户时，如果有有效期则勾选
            self.expire_check.setChecked(True)
            self.expire_edit.setDateTime(self.user.expire_time)
        elif not self.user:
            # 新建用户时，默认勾选并设置5天有效期
            self.expire_check.setChecked(True)
            self.expire_edit.setEnabled(True)
            
        expire_box.addLayout(expire_header)
        expire_box.addWidget(self.expire_edit)
        row2.addWidget(expire_container, 1)  # stretch=1
        
        # 右：最大使用次数 (使用容器控制宽度)
        usage_container = QWidget()
        usage_box = QVBoxLayout(usage_container)
        usage_box.setContentsMargins(0, 0, 0, 0)
        usage_box.setSpacing(8)
        usage_label = QLabel("最大使用次数")
        usage_label.setStyleSheet(f"color: {PREMIUM_COLORS['text_body']}; font-weight: 600; font-size: 13px;")
        
        self.usage_spin = QSpinBox()
        self.usage_spin.setRange(-1, 999999)
        self.usage_spin.setValue(-1)
        if self.user and self.user.max_usage_count is not None:
            self.usage_spin.setValue(self.user.max_usage_count)
            
        usage_hint = QLabel("填写 -1 表示不限制次数")
        usage_hint.setStyleSheet(f"color: {PREMIUM_COLORS['text_hint']}; font-size: 11px;")
            
        usage_box.addWidget(usage_label)
        usage_box.addWidget(self.usage_spin)
        usage_box.addWidget(usage_hint)
        row2.addWidget(usage_container, 1)  # stretch=1，与左边等宽
        
        form_layout.addLayout(row2)
        
        # 最大设备数 & 最大名片数 (一行两列)
        row3 = QHBoxLayout()
        row3.setSpacing(20)
        
        # 左：最大设备数
        device_container = QWidget()
        device_box = QVBoxLayout(device_container)
        device_box.setContentsMargins(0, 0, 0, 0)
        device_box.setSpacing(8)
        device_label = QLabel("最大设备数")
        device_label.setStyleSheet(f"color: {PREMIUM_COLORS['text_body']}; font-weight: 600; font-size: 13px;")
        
        self.device_spin = QSpinBox()
        self.device_spin.setRange(-1, 100)
        self.device_spin.setValue(2)  # 新用户默认 2 台设备
        if self.user:
            max_dev = getattr(self.user, 'max_device_count', 2)
            if max_dev is not None:
                self.device_spin.setValue(max_dev)
            
        device_hint = QLabel("默认2台，-1表示使用全局配置")
        device_hint.setStyleSheet(f"color: {PREMIUM_COLORS['text_hint']}; font-size: 11px;")
            
        device_box.addWidget(device_label)
        device_box.addWidget(self.device_spin)
        device_box.addWidget(device_hint)
        row3.addWidget(device_container, 1)
        
        # 右：最大名片数
        card_container = QWidget()
        card_box = QVBoxLayout(card_container)
        card_box.setContentsMargins(0, 0, 0, 0)
        card_box.setSpacing(8)
        card_label = QLabel("最大名片数")
        card_label.setStyleSheet(f"color: {PREMIUM_COLORS['text_body']}; font-weight: 600; font-size: 13px;")
        
        self.card_spin = QSpinBox()
        self.card_spin.setRange(-1, 999999)
        self.card_spin.setValue(-1)  # 新用户默认不限制
        if self.user:
            max_card = getattr(self.user, 'max_card_count', -1)
            if max_card is not None:
                self.card_spin.setValue(max_card)
            
        card_hint = QLabel("填写 -1 表示不限制")
        card_hint.setStyleSheet(f"color: {PREMIUM_COLORS['text_hint']}; font-size: 11px;")
            
        card_box.addWidget(card_label)
        card_box.addWidget(self.card_spin)
        card_box.addWidget(card_hint)
        row3.addWidget(card_container, 1)
        
        form_layout.addLayout(row3)
        
        # 状态开关
        status_layout = QHBoxLayout()
        
        status_info = QVBoxLayout()
        status_info.setSpacing(2)
        status_label = QLabel("账号状态")
        status_label.setStyleSheet(f"color: {PREMIUM_COLORS['text_body']}; font-weight: 600; font-size: 13px;")
        status_hint = QLabel("禁用后用户无法登录系统")
        status_hint.setStyleSheet(f"color: {PREMIUM_COLORS['text_hint']}; font-size: 11px;")
        status_info.addWidget(status_label)
        status_info.addWidget(status_hint)
        
        self.active_check = QCheckBox("启用此账号")
        self.active_check.setStyleSheet(self._get_check_style())
        self.active_check.setChecked(True)
        if self.user: self.active_check.setChecked(self.user.is_active)
        
        status_layout.addLayout(status_info)
        status_layout.addStretch()
        status_layout.addWidget(self.active_check)
        
        form_layout.addLayout(status_layout)
        form_layout.addStretch()
        
        layout.addWidget(form_widget)
        
        # === 3. 底部按钮 ===
        btn_layout = QHBoxLayout()
        btn_layout.setContentsMargins(32, 0, 32, 32)
        btn_layout.setSpacing(16)
        
        cancel_btn = QPushButton("取消")
        cancel_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        cancel_btn.setFixedSize(100, 44)
        cancel_btn.setStyleSheet(f"""
            QPushButton {{
                background: white;
                color: {PREMIUM_COLORS['text_body']}; 
                border: 1px solid {PREMIUM_COLORS['border']}; 
                border-radius: 22px; 
                font-weight: 600;
                font-size: 14px;
            }}
            QPushButton:hover {{ 
                background: {PREMIUM_COLORS['background']}; 
                border-color: {PREMIUM_COLORS['text_body']};
            }}
        """)
        cancel_btn.clicked.connect(self.reject)
        
        save_btn = GradientButton(
            "保存更改", 
            PREMIUM_COLORS['gradient_blue_start'], 
            PREMIUM_COLORS['gradient_blue_end']
        )
        save_btn.setFixedSize(140, 44)
        save_btn.clicked.connect(self.save_user)
        
        btn_layout.addStretch()
        btn_layout.addWidget(cancel_btn)
        btn_layout.addWidget(save_btn)
        
        layout.addLayout(btn_layout)
    
    def _create_field(self, label_text, widget, hint_text=None):
        """辅助方法：创建带标签的字段布局"""
        layout = QVBoxLayout()
        layout.setSpacing(6)
        
        label = QLabel(label_text)
        label.setStyleSheet(f"color: {PREMIUM_COLORS['text_body']}; font-weight: 600; font-size: 13px;")
        layout.addWidget(label)
        
        layout.addWidget(widget)
        
        if hint_text:
            hint = QLabel(hint_text)
            hint.setStyleSheet(f"color: {PREMIUM_COLORS['text_hint']}; font-size: 11px;")
            layout.addWidget(hint)
            
        return layout

    def _get_check_style(self):
        return f"""
            QCheckBox {{
                color: {PREMIUM_COLORS['text_heading']};
                font-size: 13px;
                spacing: 8px;
            }}
            QCheckBox::indicator {{
                width: 20px;
                height: 20px;
                border-radius: 5px;
                border: 1px solid {PREMIUM_COLORS['border']};
                background: white;
            }}
            QCheckBox::indicator:checked {{
                background: {PREMIUM_COLORS['gradient_blue_start']};
                border-color: {PREMIUM_COLORS['gradient_blue_start']};
                image: url(gui/assets/check.png); /* 这里的图标如果没有可能不显示，但颜色会变 */
            }}
        """
    
    def _setup_calendar_style(self, date_edit):
        """为日期选择器设置清晰的日历弹出样式"""
        calendar = date_edit.calendarWidget()
        if calendar:
            calendar.setStyleSheet(f"""
                /* 日历整体背景 */
                QCalendarWidget {{
                    background-color: white;
                    border: 1px solid {PREMIUM_COLORS['border']};
                    border-radius: 8px;
                }}
                
                /* 导航栏（年月选择区域） */
                QCalendarWidget QWidget#qt_calendar_navigationbar {{
                    background-color: {PREMIUM_COLORS['gradient_blue_start']};
                    border-top-left-radius: 8px;
                    border-top-right-radius: 8px;
                    padding: 8px;
                    min-height: 40px;
                }}
                
                /* 月份/年份按钮 */
                QCalendarWidget QToolButton {{
                    color: white;
                    background-color: transparent;
                    border: none;
                    border-radius: 4px;
                    padding: 6px 12px;
                    font-size: 14px;
                    font-weight: 600;
                }}
                QCalendarWidget QToolButton:hover {{
                    background-color: rgba(255, 255, 255, 0.2);
                }}
                QCalendarWidget QToolButton:pressed {{
                    background-color: rgba(255, 255, 255, 0.3);
                }}
                
                /* 左右箭头按钮 */
                QCalendarWidget QToolButton#qt_calendar_prevmonth,
                QCalendarWidget QToolButton#qt_calendar_nextmonth {{
                    qproperty-icon: none;
                    color: white;
                    font-size: 16px;
                    font-weight: bold;
                    min-width: 30px;
                }}
                QCalendarWidget QToolButton#qt_calendar_prevmonth {{
                    qproperty-text: "◀";
                }}
                QCalendarWidget QToolButton#qt_calendar_nextmonth {{
                    qproperty-text: "▶";
                }}
                
                /* 年月下拉菜单 */
                QCalendarWidget QMenu {{
                    background-color: white;
                    border: 1px solid {PREMIUM_COLORS['border']};
                    border-radius: 6px;
                    padding: 4px;
                }}
                QCalendarWidget QMenu::item {{
                    padding: 8px 20px;
                    border-radius: 4px;
                    color: {PREMIUM_COLORS['text_heading']};
                }}
                QCalendarWidget QMenu::item:selected {{
                    background-color: {PREMIUM_COLORS['primary_light']};
                    color: {PREMIUM_COLORS['gradient_blue_start']};
                }}
                
                /* 年份输入框 */
                QCalendarWidget QSpinBox {{
                    background-color: rgba(255, 255, 255, 0.2);
                    border: 1px solid rgba(255, 255, 255, 0.3);
                    border-radius: 4px;
                    color: white;
                    padding: 4px 8px;
                    font-size: 13px;
                    selection-background-color: rgba(255, 255, 255, 0.3);
                }}
                QCalendarWidget QSpinBox::up-button,
                QCalendarWidget QSpinBox::down-button {{
                    width: 16px;
                    background-color: transparent;
                    border: none;
                }}
                
                /* 星期标题行 */
                QCalendarWidget QWidget {{
                    alternate-background-color: #f8fafc;
                }}
                
                /* 日期表格 */
                QCalendarWidget QAbstractItemView:enabled {{
                    background-color: white;
                    color: {PREMIUM_COLORS['text_heading']};
                    selection-background-color: {PREMIUM_COLORS['gradient_blue_start']};
                    selection-color: white;
                    outline: none;
                    font-size: 13px;
                }}
                
                /* 日期单元格 */
                QCalendarWidget QAbstractItemView:enabled {{
                    border: none;
                }}
                
                /* 星期标题 */
                QCalendarWidget QHeaderView::section {{
                    background-color: #f1f5f9;
                    color: {PREMIUM_COLORS['text_body']};
                    font-weight: 600;
                    font-size: 11px;
                    padding: 6px;
                    border: none;
                }}
                
                /* 周末颜色 (周六周日) */
                QCalendarWidget QAbstractItemView:enabled {{
                    color: {PREMIUM_COLORS['text_heading']};
                }}
            """)
        
    def save_user(self):
        username = self.username_input.text().strip()
        password = self.password_input.text().strip()
        role = self.role_combo.currentText()
        is_active = self.active_check.isChecked()
        
        expire_time = None
        if self.expire_check.isChecked():
            expire_time = self.expire_edit.dateTime().toPyDateTime()
            
        max_usage_count = self.usage_spin.value()
        max_device_count = self.device_spin.value()
        max_card_count = self.card_spin.value()
        
        if not self.user and not username:
            QMessageBox.warning(self, "提示", "请输入用户名")
            return
            
        if not self.user and not password:
            QMessageBox.warning(self, "提示", "请输入密码")
            return
            
        try:
            if self.user:
                DatabaseManager.update_user(
                    str(self.user.id),
                    password if password else None,
                    role,
                    is_active,
                    expire_time=expire_time,
                    max_usage_count=max_usage_count,
                    max_device_count=max_device_count,
                    max_card_count=max_card_count
                )
            else:
                DatabaseManager.create_user(
                    username, 
                    password, 
                    role, 
                    is_active,
                    expire_time=expire_time,
                    max_usage_count=max_usage_count,
                    max_device_count=max_device_count,
                    max_card_count=max_card_count
                )
            
            self.accept()
        except ValueError as e:
            QMessageBox.warning(self, "错误", str(e))
        except Exception as e:
            QMessageBox.critical(self, "错误", f"操作失败: {str(e)}")


class DeviceListDialog(QDialog):
    """设备列表对话框 - 重新设计版"""
    
    def __init__(self, parent=None, user=None):
        super().__init__(parent)
        self.user = user
        self.setWindowTitle(f"设备管理 - {user.username}")
        self.setFixedSize(800, 500)
        self.setStyleSheet("QDialog { background-color: white; }")
        self.init_ui()
        
    def init_ui(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        self.setLayout(layout)
        
        # === 1. 顶部 Header ===
        header = QFrame()
        header.setFixedHeight(90)
        header.setStyleSheet(f"""
            QFrame {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 {PREMIUM_COLORS['gradient_blue_start']}, 
                    stop:1 {PREMIUM_COLORS['gradient_blue_end']});
            }}
        """)
        
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(32, 0, 32, 0)
        header_layout.setSpacing(16)
        
        # 标题文字
        title_info = QVBoxLayout()
        title_info.setSpacing(4)
        title_info.setAlignment(Qt.AlignmentFlag.AlignVCenter)
        
        title_lbl = QLabel(f"{self.user.username} 的设备")
        title_lbl.setStyleSheet("color: white; font-size: 24px; font-weight: 800;")
        
        subtitle_lbl = QLabel("管理已授权的登录设备，移除后该设备需重新登录")
        subtitle_lbl.setStyleSheet("color: rgba(255,255,255,0.8); font-size: 12px; font-weight: 500;")
        
        title_info.addWidget(title_lbl)
        title_info.addWidget(subtitle_lbl)
        
        # 图标
        icon_bg = QLabel("📱")
        icon_bg.setFixedSize(48, 48)
        icon_bg.setAlignment(Qt.AlignmentFlag.AlignCenter)
        icon_bg.setStyleSheet("""
            background: rgba(255,255,255,0.2);
            border-radius: 24px;
            font-size: 24px;
        """)
        
        header_layout.addLayout(title_info)
        header_layout.addWidget(icon_bg)
        
        layout.addWidget(header)
        
        # === 2. 表格区域 ===
        table_container = QWidget()
        table_layout = QVBoxLayout(table_container)
        table_layout.setContentsMargins(20, 20, 20, 20)
        
        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(['设备名称', '类型', '设备ID', '最近活动', '操作'])
        self.table.verticalHeader().setVisible(False)
        self.table.setShowGrid(False)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.table.setAlternatingRowColors(True)
        
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed)
        
        self.table.setColumnWidth(1, 100) # 类型
        self.table.setColumnWidth(3, 140) # 时间
        self.table.setColumnWidth(4, 80)  # 操作
        
        self.table.setStyleSheet(f"""
            QTableWidget {{
                background: white;
                border: 1px solid {PREMIUM_COLORS['border_light']};
                border-radius: 8px;
                selection-background-color: {PREMIUM_COLORS['primary_light']};
                alternate-background-color: #f8fafc;
            }}
            QTableWidget::item {{
                padding: 12px;
                border-bottom: 1px solid {PREMIUM_COLORS['border_light']};
                color: {PREMIUM_COLORS['text_body']};
            }}
            QHeaderView::section {{
                background: #f1f5f9;
                border: none;
                border-bottom: 1px solid {PREMIUM_COLORS['border_light']};
                padding: 12px;
                color: {PREMIUM_COLORS['text_heading']};
                font-weight: 700;
                font-size: 12px;
            }}
        """)
        
        table_layout.addWidget(self.table)
        layout.addWidget(table_container, 1)
        
        self.load_devices()
        
        # === 3. 底部按钮 ===
        btn_layout = QHBoxLayout()
        btn_layout.setContentsMargins(20, 0, 20, 20)
        btn_layout.addStretch()
        
        close_btn = QPushButton("关闭")
        close_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        close_btn.setFixedSize(100, 36)
        close_btn.setStyleSheet(f"""
            QPushButton {{
                background: white;
                color: {PREMIUM_COLORS['text_body']};
                border: 1px solid {PREMIUM_COLORS['border']};
                border-radius: 18px;
                font-weight: 600;
                font-size: 13px;
            }}
            QPushButton:hover {{
                background: {PREMIUM_COLORS['background']};
                border-color: {PREMIUM_COLORS['text_body']};
            }}
        """)
        close_btn.clicked.connect(self.accept)
        btn_layout.addWidget(close_btn)
        
        layout.addLayout(btn_layout)
        
    def load_devices(self):
        devices = Device.objects(user=self.user).order_by('-last_login')
        self.table.setRowCount(len(devices))
        
        for row, device in enumerate(devices):
            self.table.setRowHeight(row, 60)
            
            # 1. 设备名称 (图标 + 名字)
            name_widget = QWidget()
            name_layout = QHBoxLayout(name_widget)
            name_layout.setContentsMargins(8, 0, 8, 0)
            name_layout.setSpacing(10)
            name_layout.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
            
            icon_lbl = QLabel("💻")
            icon_lbl.setStyleSheet("font-size: 18px;")
            
            name_lbl = QLabel(device.device_name)
            name_lbl.setStyleSheet(f"font-weight: 600; color: {PREMIUM_COLORS['text_heading']}; font-size: 13px;")
            
            name_layout.addWidget(icon_lbl)
            name_layout.addWidget(name_lbl)
            self.table.setCellWidget(row, 0, name_widget)
            
            # 2. 类型
            type_item = QTableWidgetItem(device.device_type or '-')
            type_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row, 1, type_item)
            
            # 3. 设备ID (截断显示)
            id_str = device.device_id
            if len(id_str) > 20:
                id_str = id_str[:8] + "..." + id_str[-8:]
            
            id_item = QTableWidgetItem(id_str)
            id_item.setToolTip(device.device_id) # 悬停显示完整ID
            id_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            id_item.setForeground(QBrush(QColor(PREMIUM_COLORS['text_hint'])))
            self.table.setItem(row, 2, id_item)
            
            # 4. 最近活动
            last_login = device.last_login.strftime('%Y-%m-%d %H:%M') if device.last_login else '-'
            time_item = QTableWidgetItem(last_login)
            time_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row, 3, time_item)
            
            # 5. 操作按钮
            action_widget = QWidget()
            action_layout = QHBoxLayout(action_widget)
            action_layout.setContentsMargins(0, 0, 0, 0)
            action_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            
            remove_btn = QPushButton("移除")
            remove_btn.setCursor(Qt.CursorShape.PointingHandCursor)
            remove_btn.setFixedSize(50, 28)
            remove_btn.setStyleSheet(f"""
                QPushButton {{
                    background: {PREMIUM_COLORS['coral']}15;
                    color: {PREMIUM_COLORS['coral']};
                    border: 1px solid {PREMIUM_COLORS['coral']}40;
                    border-radius: 4px;
                    font-size: 12px;
                    font-weight: 600;
                }}
                QPushButton:hover {{
                    background: {PREMIUM_COLORS['coral']};
                    color: white;
                    border-color: {PREMIUM_COLORS['coral']};
                }}
            """)
            # 使用闭包捕获当前行的设备对象
            remove_btn.clicked.connect(lambda checked, d=device: self.remove_device(d))
            
            action_layout.addWidget(remove_btn)
            self.table.setCellWidget(row, 4, action_widget)
    
    def remove_device(self, device):
        """移除设备"""
        reply = QMessageBox.question(
            self,
            "确认移除",
            f"确定要移除设备 {device.device_name} 吗？\n\n该设备将需要重新登录才能使用。",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            try:
                device.delete()
                self.load_devices()
                QMessageBox.information(self, "成功", "设备已移除")
            except Exception as e:
                QMessageBox.critical(self, "错误", f"移除设备失败：{str(e)}")


class CompactStatWidget(QFrame):
    """紧凑型统计组件"""
    def __init__(self, title, value, icon, color_start, color_end, parent=None):
        super().__init__(parent)
        self.value = value
        self._setup_ui(title, icon, color_start, color_end)
        
    def _setup_ui(self, title, icon, color_start, color_end):
        self.setFixedSize(140, 50)
        layout = QHBoxLayout(self)
        layout.setContentsMargins(10, 5, 10, 5)
        layout.setSpacing(8)
        
        # 背景样式
        self.setStyleSheet(f"""
            CompactStatWidget {{
                background: white;
                border-radius: 12px;
                border: 1px solid {PREMIUM_COLORS['border_light']};
            }}
        """)
        
        # 图标
        icon_lbl = QLabel(icon)
        icon_lbl.setFixedSize(32, 32)
        icon_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        icon_lbl.setStyleSheet(f"""
            background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 {color_start}, stop:1 {color_end});
            color: white;
            border-radius: 8px;
            font-size: 16px;
        """)
        layout.addWidget(icon_lbl)
        
        # 文本
        text_layout = QVBoxLayout()
        text_layout.setSpacing(0)
        text_layout.setAlignment(Qt.AlignmentFlag.AlignVCenter)
        
        self.value_lbl = QLabel(str(self.value))
        self.value_lbl.setStyleSheet(f"font-size: 16px; font-weight: 800; color: {PREMIUM_COLORS['text_heading']};")
        text_layout.addWidget(self.value_lbl)
        
        title_lbl = QLabel(title)
        title_lbl.setStyleSheet(f"font-size: 10px; color: {PREMIUM_COLORS['text_hint']};")
        text_layout.addWidget(title_lbl)
        
        layout.addLayout(text_layout)
        
    def update_value(self, value):
        self.value = value
        self.value_lbl.setText(str(value))


# ========== 自定义用户列表组件 ==========

# 列宽配置 (固定宽度，确保对齐)
USER_LIST_COLUMNS = {
    'avatar': 50,
    'user': 140,
    'role': 70,
    'device': 70,
    'cards': 80,
    'usage': 100,
    'expire': 110,
    'status': 60,
    'activity': 100,
    'actions': 180,
}


class UserListHeader(QFrame):
    """用户列表表头"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFixedHeight(44)
        self.setStyleSheet(f"""
            UserListHeader {{
                background: {PREMIUM_COLORS['background']};
                border: none;
                border-bottom: 1px solid {PREMIUM_COLORS['border_light']};
            }}
        """)
        self._setup_ui()
    
    def _setup_ui(self):
        layout = QHBoxLayout(self)
        layout.setContentsMargins(16, 0, 16, 0)
        layout.setSpacing(0)
        
        headers = [
            ('', USER_LIST_COLUMNS['avatar']),
            ('用户', USER_LIST_COLUMNS['user']),
            ('角色', USER_LIST_COLUMNS['role']),
            ('设备', USER_LIST_COLUMNS['device']),
            ('名片', USER_LIST_COLUMNS['cards']),
            ('额度', USER_LIST_COLUMNS['usage']),
            ('有效期', USER_LIST_COLUMNS['expire']),
            ('状态', USER_LIST_COLUMNS['status']),
            ('最近活动', USER_LIST_COLUMNS['activity']),
            ('操作', USER_LIST_COLUMNS['actions']),
        ]
        
        for text, width in headers:
            lbl = QLabel(text)
            lbl.setFixedWidth(width)
            lbl.setStyleSheet(f"""
                color: {PREMIUM_COLORS['text_hint']};
                font-size: 12px;
                font-weight: 700;
                text-transform: uppercase;
                padding-left: 4px;
            """)
            layout.addWidget(lbl)
        
        layout.addStretch()


class UserRowWidget(QFrame):
    """单行用户数据组件"""
    
    # 定义信号
    edit_clicked = pyqtSignal(object)
    device_clicked = pyqtSignal(object)
    enter_clicked = pyqtSignal(object)
    toggle_clicked = pyqtSignal(object)
    delete_clicked = pyqtSignal(object)
    import_cards_clicked = pyqtSignal(object)
    
    def __init__(self, user, device_count, card_count=0, parent=None):
        super().__init__(parent)
        self.user = user
        self.device_count = device_count
        self.card_count = card_count
        self.setFixedHeight(64)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self._is_hovered = False
        self._setup_ui()
        
    def _setup_ui(self):
        self.setStyleSheet(f"""
            UserRowWidget {{
                background: white;
                border: none;
                border-bottom: 1px solid {PREMIUM_COLORS['border_light']};
            }}
            UserRowWidget:hover {{
                background: #fafbfc;
            }}
        """)
        
        layout = QHBoxLayout(self)
        layout.setContentsMargins(16, 8, 16, 8)
        layout.setSpacing(0)
        
        # 1. 头像
        self._add_avatar(layout)
        
        # 2. 用户信息
        self._add_user_info(layout)
        
        # 3. 角色
        self._add_role(layout)
        
        # 4. 设备
        self._add_device(layout)
        
        # 5. 名片
        self._add_cards(layout)
        
        # 6. 使用额度
        self._add_usage(layout)
        
        # 7. 有效期
        self._add_expire(layout)
        
        # 8. 状态
        self._add_status(layout)
        
        # 9. 最近活动
        self._add_activity(layout)
        
        # 10. 操作按钮
        self._add_actions(layout)
        
        layout.addStretch()
    
    def _add_avatar(self, layout):
        """添加头像"""
        container = QWidget()
        container.setFixedWidth(USER_LIST_COLUMNS['avatar'])
        c_layout = QHBoxLayout(container)
        c_layout.setContentsMargins(0, 0, 8, 0)
        c_layout.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        
        avatar = QLabel(self.user.username[0].upper())
        avatar.setFixedSize(36, 36)
        avatar.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        # 智能颜色生成
        if self.user.role == 'admin':
            bg_gradient = f"qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 {PREMIUM_COLORS['gradient_blue_start']}, stop:1 {PREMIUM_COLORS['gradient_blue_end']})"
        else:
            colors = [
                (PREMIUM_COLORS['gradient_green_start'], PREMIUM_COLORS['gradient_green_end']),
                (PREMIUM_COLORS['gradient_orange_start'], PREMIUM_COLORS['gradient_orange_end']),
                (PREMIUM_COLORS['gradient_purple_start'], PREMIUM_COLORS['gradient_purple_end']),
            ]
            c_start, c_end = colors[sum(ord(c) for c in self.user.username) % len(colors)]
            bg_gradient = f"qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 {c_start}, stop:1 {c_end})"
        
        avatar.setStyleSheet(f"""
            background: {bg_gradient};
            color: white;
            border-radius: 18px;
            font-size: 15px;
            font-weight: 700;
        """)
        
        c_layout.addWidget(avatar)
        layout.addWidget(container)
    
    def _add_user_info(self, layout):
        """添加用户信息"""
        container = QWidget()
        container.setFixedWidth(USER_LIST_COLUMNS['user'])
        c_layout = QVBoxLayout(container)
        c_layout.setContentsMargins(4, 0, 4, 0)
        c_layout.setSpacing(2)
        c_layout.setAlignment(Qt.AlignmentFlag.AlignVCenter)
        
        name_label = QLabel(self.user.username)
        name_label.setStyleSheet(f"""
            font-size: 14px;
            font-weight: 600;
            color: {PREMIUM_COLORS['text_heading']};
        """)
        c_layout.addWidget(name_label)
        
        created_str = self.user.created_at.strftime('%Y-%m-%d') if self.user.created_at else '未知'
        created_label = QLabel(f"加入: {created_str}")
        created_label.setStyleSheet(f"""
            font-size: 11px;
            color: {PREMIUM_COLORS['text_hint']};
        """)
        c_layout.addWidget(created_label)
        
        layout.addWidget(container)
    
    def _add_role(self, layout):
        """添加角色"""
        container = QWidget()
        container.setFixedWidth(USER_LIST_COLUMNS['role'])
        c_layout = QHBoxLayout(container)
        c_layout.setContentsMargins(0, 0, 4, 0)
        c_layout.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        
        role_label = QLabel("管理员" if self.user.role == 'admin' else "用户")
        role_label.setFixedHeight(24)
        
        if self.user.role == 'admin':
            role_label.setStyleSheet(f"""
                background: {PREMIUM_COLORS['gradient_blue_start']}18;
                color: {PREMIUM_COLORS['gradient_blue_start']};
                border: 1px solid {PREMIUM_COLORS['gradient_blue_start']}40;
                border-radius: 12px;
                padding: 2px 10px;
                font-size: 11px;
                font-weight: 600;
            """)
        else:
            role_label.setStyleSheet(f"""
                background: {PREMIUM_COLORS['text_hint']}15;
                color: {PREMIUM_COLORS['text_body']};
                border: 1px solid {PREMIUM_COLORS['text_hint']}35;
                border-radius: 12px;
                padding: 2px 10px;
                font-size: 11px;
                font-weight: 500;
            """)
        
        c_layout.addWidget(role_label)
        layout.addWidget(container)
    
    def _add_device(self, layout):
        """添加设备数 - 显示 当前数/限制数"""
        container = QWidget()
        container.setFixedWidth(USER_LIST_COLUMNS['device'])
        c_layout = QHBoxLayout(container)
        c_layout.setContentsMargins(0, 0, 4, 0)
        c_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        c_layout.setSpacing(4)
        
        d_icon = QLabel("💻")
        d_icon.setStyleSheet("font-size: 14px;")
        
        # 获取最大设备数限制
        max_device = getattr(self.user, 'max_device_count', 2)
        if max_device == -1:
            # -1 表示使用全局配置，显示为无限
            d_text = QLabel(f"{self.device_count}")
        else:
            d_text = QLabel(f"{self.device_count}/{max_device}")
        
        d_text.setStyleSheet(f"font-weight: 600; color: {PREMIUM_COLORS['text_heading']}; font-size: 12px;")
        
        c_layout.addWidget(d_icon)
        c_layout.addWidget(d_text)
        layout.addWidget(container)
    
    def _add_cards(self, layout):
        """添加名片数 - 显示 当前数/限制数"""
        container = QWidget()
        container.setFixedWidth(USER_LIST_COLUMNS['cards'])
        c_layout = QHBoxLayout(container)
        c_layout.setContentsMargins(0, 0, 4, 0)
        c_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        c_layout.setSpacing(4)
        
        c_icon = QLabel("📇")
        c_icon.setStyleSheet("font-size: 14px;")
        
        # 获取最大名片数限制
        max_card = getattr(self.user, 'max_card_count', -1)
        if max_card == -1:
            # -1 表示无限制
            c_text = QLabel(f"∞ ({self.card_count})")
        else:
            c_text = QLabel(f"{self.card_count}/{max_card}")
        
        c_text.setStyleSheet(f"font-weight: 600; color: {PREMIUM_COLORS['text_heading']}; font-size: 12px;")
        
        c_layout.addWidget(c_icon)
        c_layout.addWidget(c_text)
        layout.addWidget(container)
    
    def _add_usage(self, layout):
        """添加使用额度"""
        container = QWidget()
        container.setFixedWidth(USER_LIST_COLUMNS['usage'])
        c_layout = QVBoxLayout(container)
        c_layout.setContentsMargins(0, 0, 8, 0)
        c_layout.setSpacing(4)
        c_layout.setAlignment(Qt.AlignmentFlag.AlignVCenter)
        
        usage_count = self.user.usage_count or 0
        max_count = self.user.max_usage_count if self.user.max_usage_count is not None else -1
        
        if max_count == -1:
            progress_percent = 0
            text = f"∞ ({usage_count})"
        else:
            progress_percent = min(100, int(usage_count / max_count * 100)) if max_count > 0 else 0
            text = f"{usage_count} / {max_count}"
        
        label = QLabel(text)
        label.setStyleSheet(f"font-size: 12px; color: {PREMIUM_COLORS['text_body']}; font-weight: 500;")
        c_layout.addWidget(label)
        
        if max_count != -1:
            # 进度条背景
            prog_bg = QFrame()
            prog_bg.setFixedSize(80, 4)
            prog_bg.setStyleSheet(f"background: {PREMIUM_COLORS['background']}; border-radius: 2px;")
            
            # 进度填充
            fill = QFrame(prog_bg)
            fill.setFixedHeight(4)
            width = int(80 * progress_percent / 100)
            fill.setFixedWidth(max(4, width))
            fill.move(0, 0)
            
            if progress_percent > 90:
                color = PREMIUM_COLORS['coral']
            elif progress_percent > 70:
                color = PREMIUM_COLORS['gradient_gold_start']
            else:
                color = PREMIUM_COLORS['gradient_green_start']
            
            fill.setStyleSheet(f"background: {color}; border-radius: 2px;")
            c_layout.addWidget(prog_bg)
        
        layout.addWidget(container)
    
    def _add_expire(self, layout):
        """添加有效期"""
        container = QWidget()
        container.setFixedWidth(USER_LIST_COLUMNS['expire'])
        c_layout = QVBoxLayout(container)
        c_layout.setContentsMargins(0, 0, 4, 0)
        c_layout.setSpacing(2)
        c_layout.setAlignment(Qt.AlignmentFlag.AlignVCenter)
        
        if self.user.expire_time:
            days = (self.user.expire_time - datetime.datetime.now()).days
            date_str = self.user.expire_time.strftime('%Y-%m-%d')
            
            date_lbl = QLabel(date_str)
            date_lbl.setStyleSheet(f"font-size: 12px; color: {PREMIUM_COLORS['text_heading']}; font-weight: 500;")
            
            status_lbl = QLabel()
            if days < 0:
                status_lbl.setText("已过期")
                status_lbl.setStyleSheet(f"color: {PREMIUM_COLORS['coral']}; font-size: 10px; font-weight: 600;")
            elif days <= 7:
                status_lbl.setText(f"剩 {days} 天")
                status_lbl.setStyleSheet(f"color: {PREMIUM_COLORS['gradient_gold_start']}; font-size: 10px; font-weight: 600;")
            else:
                status_lbl.setText(f"剩 {days} 天")
                status_lbl.setStyleSheet(f"color: {PREMIUM_COLORS['text_hint']}; font-size: 10px;")
            
            c_layout.addWidget(date_lbl)
            c_layout.addWidget(status_lbl)
        else:
            lbl = QLabel("永久有效")
            lbl.setStyleSheet(f"color: {PREMIUM_COLORS['mint']}; font-weight: 600; font-size: 12px;")
            c_layout.addWidget(lbl)
        
        layout.addWidget(container)
    
    def _add_status(self, layout):
        """添加状态"""
        container = QWidget()
        container.setFixedWidth(USER_LIST_COLUMNS['status'])
        c_layout = QHBoxLayout(container)
        c_layout.setContentsMargins(0, 0, 4, 0)
        c_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        status_badge = QLabel("正常" if self.user.is_active else "禁用")
        if self.user.is_active:
            status_badge.setStyleSheet(f"""
                background: {PREMIUM_COLORS['gradient_green_start']}15;
                color: {PREMIUM_COLORS['gradient_green_start']};
                padding: 3px 10px;
                border-radius: 10px;
                font-size: 11px;
                font-weight: 600;
            """)
        else:
            status_badge.setStyleSheet(f"""
                background: {PREMIUM_COLORS['coral']}15;
                color: {PREMIUM_COLORS['coral']};
                padding: 3px 10px;
                border-radius: 10px;
                font-size: 11px;
                font-weight: 600;
            """)
        
        c_layout.addWidget(status_badge)
        layout.addWidget(container)
    
    def _add_activity(self, layout):
        """添加最近活动"""
        container = QWidget()
        container.setFixedWidth(USER_LIST_COLUMNS['activity'])
        c_layout = QVBoxLayout(container)
        c_layout.setContentsMargins(0, 0, 4, 0)
        c_layout.setAlignment(Qt.AlignmentFlag.AlignVCenter)
        
        if self.user.last_login:
            t_str = self.user.last_login.strftime('%m-%d %H:%M')
            l1 = QLabel(t_str)
            l1.setStyleSheet(f"color: {PREMIUM_COLORS['text_body']}; font-size: 12px;")
            c_layout.addWidget(l1)
        else:
            l = QLabel("从未登录")
            l.setStyleSheet(f"color: {PREMIUM_COLORS['text_hint']}; font-size: 12px;")
            c_layout.addWidget(l)
        
        layout.addWidget(container)
    
    def _add_actions(self, layout):
        """添加操作按钮"""
        container = QWidget()
        container.setFixedWidth(USER_LIST_COLUMNS['actions'])
        c_layout = QHBoxLayout(container)
        c_layout.setContentsMargins(0, 0, 0, 0)
        c_layout.setSpacing(6)
        c_layout.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        
        # 辅助函数创建操作按钮
        def create_op_btn(text, color):
            btn = QPushButton(text)
            btn.setFixedSize(44, 26)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.setStyleSheet(f"""
                QPushButton {{
                    background: transparent;
                    color: {color};
                    border: 1px solid {color}40;
                    border-radius: 4px;
                    font-size: 11px;
                    font-weight: 600;
                }}
                QPushButton:hover {{
                    background: {color}12;
                    border-color: {color};
                }}
            """)
            return btn
        
        # 进入用户端
        btn_enter = create_op_btn("进入", PREMIUM_COLORS['mint'])
        btn_enter.clicked.connect(lambda: self.enter_clicked.emit(self.user))
        c_layout.addWidget(btn_enter)
        
        # 设备
        btn_dev = create_op_btn("设备", PREMIUM_COLORS['text_body'])
        btn_dev.clicked.connect(lambda: self.device_clicked.emit(self.user))
        c_layout.addWidget(btn_dev)
        
        # 编辑
        btn_edit = create_op_btn("编辑", PREMIUM_COLORS['gradient_blue_start'])
        btn_edit.clicked.connect(lambda: self.edit_clicked.emit(self.user))
        c_layout.addWidget(btn_edit)
        
        # 更多
        more_btn = QPushButton("•••")
        more_btn.setFixedSize(26, 26)
        more_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        more_btn.setStyleSheet(f"""
            QPushButton {{
                background: transparent;
                color: {PREMIUM_COLORS['text_hint']};
                border: none;
                font-size: 12px;
                font-weight: 900;
                border-radius: 13px;
            }}
            QPushButton:hover {{
                background: {PREMIUM_COLORS['background']};
                color: {PREMIUM_COLORS['text_body']};
            }}
        """)
        more_btn.clicked.connect(lambda: self._show_more_menu(more_btn))
        c_layout.addWidget(more_btn)
        
        layout.addWidget(container)
    
    def _show_more_menu(self, button):
        """显示更多操作菜单"""
        from PyQt6.QtWidgets import QMenu
        
        menu = QMenu(self)
        menu.setWindowFlags(menu.windowFlags() | Qt.WindowType.FramelessWindowHint)
        menu.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        
        menu.setStyleSheet(f"""
            QMenu {{
                background: white;
                border: 1px solid {PREMIUM_COLORS['border_light']};
                border-radius: 10px;
                padding: 6px;
            }}
            QMenu::item {{
                padding: 8px 20px;
                border-radius: 6px;
                color: {PREMIUM_COLORS['text_body']};
                font-size: 12px;
                font-weight: 500;
            }}
            QMenu::item:selected {{
                background: {PREMIUM_COLORS['background']};
                color: {PREMIUM_COLORS['gradient_blue_start']};
            }}
            QMenu::separator {{
                height: 1px;
                background: {PREMIUM_COLORS['border_light']};
                margin: 4px 0;
            }}
        """)
        
        # 添加阴影
        shadow = QGraphicsDropShadowEffect(menu)
        shadow.setBlurRadius(16)
        shadow.setColor(QColor(0, 0, 0, 25))
        shadow.setOffset(0, 4)
        menu.setGraphicsEffect(shadow)
        
        # 导入名片
        import_action = menu.addAction("📥 导入名片")
        import_action.triggered.connect(lambda: self.import_cards_clicked.emit(self.user))
        
        # 禁用/启用
        toggle_action = menu.addAction("🚫 禁用账号" if self.user.is_active else "✅ 启用账号")
        toggle_action.triggered.connect(lambda: self.toggle_clicked.emit(self.user))
        
        menu.addSeparator()
        
        # 删除
        if self.user.username != 'admin':
            delete_action = menu.addAction("🗑️ 删除用户")
            delete_action.triggered.connect(lambda: self.delete_clicked.emit(self.user))
        
        menu.exec(button.mapToGlobal(button.rect().bottomLeft()))


class UserListWidget(QWidget):
    """自定义用户列表组件 (替代 QTableWidget)"""
    
    # 定义信号，向外传递用户操作
    edit_user = pyqtSignal(object)
    device_user = pyqtSignal(object)
    enter_user = pyqtSignal(object)
    toggle_user = pyqtSignal(object)
    delete_user = pyqtSignal(object)
    import_cards_user = pyqtSignal(object)
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.row_widgets = []
        self._setup_ui()
    
    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        
        # 表头
        self.header = UserListHeader()
        layout.addWidget(self.header)
        
        # 滚动区域
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.scroll_area.setStyleSheet(f"""
            QScrollArea {{
                border: none;
                background: transparent;
            }}
            QScrollBar:vertical {{
                background: transparent;
                width: 8px;
                margin: 0;
            }}
            QScrollBar::handle:vertical {{
                background: {PREMIUM_COLORS['border']};
                border-radius: 4px;
                min-height: 30px;
            }}
            QScrollBar::handle:vertical:hover {{
                background: {PREMIUM_COLORS['text_hint']};
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                height: 0;
            }}
            QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {{
                background: transparent;
            }}
        """)
        
        # 内容容器
        self.content_widget = QWidget()
        self.content_widget.setStyleSheet("background: white;")
        self.content_layout = QVBoxLayout(self.content_widget)
        self.content_layout.setContentsMargins(0, 0, 0, 0)
        self.content_layout.setSpacing(0)
        self.content_layout.setAlignment(Qt.AlignmentFlag.AlignTop)
        
        self.scroll_area.setWidget(self.content_widget)
        layout.addWidget(self.scroll_area, 1)
    
    def set_users(self, users, device_counts, card_counts=None):
        """设置用户列表数据"""
        if card_counts is None:
            card_counts = {}
        
        # 清空现有行
        for widget in self.row_widgets:
            widget.deleteLater()
        self.row_widgets.clear()
        
        # 显示空状态
        if not users:
            empty_label = QLabel("暂无用户数据")
            empty_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            empty_label.setStyleSheet(f"""
                color: {PREMIUM_COLORS['text_hint']};
                font-size: 14px;
                padding: 60px;
            """)
            self.content_layout.addWidget(empty_label)
            self.row_widgets.append(empty_label)
            return
        
        # 添加用户行
        for user in users:
            device_count = device_counts.get(str(user.id), 0)
            card_count = card_counts.get(str(user.id), 0)
            row = UserRowWidget(user, device_count, card_count)
            
            # 连接信号
            row.edit_clicked.connect(self.edit_user.emit)
            row.device_clicked.connect(self.device_user.emit)
            row.enter_clicked.connect(self.enter_user.emit)
            row.toggle_clicked.connect(self.toggle_user.emit)
            row.delete_clicked.connect(self.delete_user.emit)
            row.import_cards_clicked.connect(self.import_cards_user.emit)
            
            self.content_layout.addWidget(row)
            self.row_widgets.append(row)


class UserManagementWidget(QWidget):
    """用户管理页面组件 - 极简布局版"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.db_manager = DatabaseManager()
        self.current_page = 1
        self.page_size = 15
        self.total_users = 0
        self.total_pages = 1
        self.stat_cards = {}
        self.init_ui()
        
    def init_ui(self):
        # 主背景
        self.setStyleSheet(f"""
            QWidget {{
                background: qlineargradient(x1:0, y1:0, x2:0.5, y2:1,
                    stop:0 #f8fafc, 
                    stop:0.6 #f1f5f9,
                    stop:1 #eef2f7);
            }}
        """)
        
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(24, 24, 24, 24)
        main_layout.setSpacing(16)
        
        # === 顶部区域 (标题 + 统计 + 操作) ===
        self._create_header(main_layout)
        
        # === 主表格区域 (包含工具栏和分页) ===
        self._create_main_card(main_layout)
        
        # 加载数据
        self.load_users()
    
    def _create_header(self, layout):
        """创建顶部区域：标题、统计、添加按钮"""
        header_layout = QHBoxLayout()
        header_layout.setSpacing(16)
        
        # 1. 标题
        title_label = QLabel("用户管理中心")
        title_label.setStyleSheet(f"""
            font-size: 24px; 
            font-weight: 800; 
            color: {PREMIUM_COLORS['text_heading']};
            background: transparent;
        """)
        header_layout.addWidget(title_label)
        
        header_layout.addSpacing(16)
        
        # 2. 统计组件 (紧凑型)
        cards_data = [
            ("总用户数", 0, "👤", PREMIUM_COLORS['gradient_blue_start'], PREMIUM_COLORS['gradient_blue_end']),
            ("活跃用户", 0, "✅", PREMIUM_COLORS['gradient_green_start'], PREMIUM_COLORS['gradient_green_end']),
            ("管理员", 0, "👑", PREMIUM_COLORS['gradient_gold_start'], PREMIUM_COLORS['gradient_gold_end']),
        ]
        
        for title, value, icon, start, end in cards_data:
            card = CompactStatWidget(title, value, icon, start, end)
            self.stat_cards[title] = card
            header_layout.addWidget(card)
            
        header_layout.addStretch()
        
        # 3. 添加按钮
        add_btn = GradientButton(
            "+ 添加用户",
            PREMIUM_COLORS['gradient_blue_start'],
            PREMIUM_COLORS['gradient_blue_end']
        )
        add_btn.setFixedSize(120, 40)
        add_btn.setStyleSheet(add_btn.styleSheet() + """
            QPushButton {
                font-size: 13px;
                border-radius: 20px;
                padding: 0 16px;
            }
        """)
        add_btn.clicked.connect(self.show_add_dialog)
        header_layout.addWidget(add_btn)
        
        layout.addLayout(header_layout)
        
    def _create_main_card(self, layout):
        """创建主内容卡片：工具栏 + 用户列表 + 分页"""
        card = GlassFrame(opacity=1.0, radius=16)
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(0, 0, 0, 0)
        card_layout.setSpacing(0)
        
        # 1. 工具栏 (搜索 + 刷新)
        toolbar = QFrame()
        toolbar.setFixedHeight(56)
        toolbar.setStyleSheet(f"border-bottom: 1px solid {PREMIUM_COLORS['border_light']}; background: transparent;")
        toolbar_layout = QHBoxLayout(toolbar)
        toolbar_layout.setContentsMargins(16, 0, 16, 0)
        
        # 搜索框
        search_container = QFrame()
        search_container.setFixedSize(260, 36)
        search_container.setStyleSheet(f"""
            QFrame {{
                background: {PREMIUM_COLORS['background']};
                border-radius: 8px;
                border: 1px solid transparent;
            }}
            QFrame:hover {{
                background: white;
                border-color: {PREMIUM_COLORS['border']};
            }}
        """)
        search_layout = QHBoxLayout(search_container)
        search_layout.setContentsMargins(10, 0, 10, 0)
        search_layout.setSpacing(8)
        
        search_icon = QLabel("🔍")
        search_icon.setStyleSheet("font-size: 14px; color: #a0aec0; border: none; background: transparent;")
        search_layout.addWidget(search_icon)
        
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("搜索用户名...")
        self.search_input.setStyleSheet(f"""
            QLineEdit {{
                border: none;
                background: transparent;
                font-size: 13px;
                color: {PREMIUM_COLORS['text_heading']};
                padding: 0;
            }}
        """)
        self.search_input.textChanged.connect(self.on_search)
        search_layout.addWidget(self.search_input)
        toolbar_layout.addWidget(search_container)
        
        toolbar_layout.addStretch()
        
        # 刷新按钮
        refresh_btn = QPushButton("刷新")
        refresh_btn.setIcon(Icons.refresh())
        refresh_btn.setFixedSize(80, 36)
        refresh_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        refresh_btn.setStyleSheet(f"""
            QPushButton {{
                background: transparent;
                border: 1px solid {PREMIUM_COLORS['border']};
                border-radius: 8px;
                color: {PREMIUM_COLORS['text_body']};
                font-weight: 600;
                font-size: 12px;
            }}
            QPushButton:hover {{
                background: {PREMIUM_COLORS['background']};
                color: {PREMIUM_COLORS['gradient_blue_start']};
                border-color: {PREMIUM_COLORS['gradient_blue_start']};
            }}
        """)
        refresh_btn.clicked.connect(self.refresh_data)
        toolbar_layout.addWidget(refresh_btn)
        
        card_layout.addWidget(toolbar)
        
        # 2. 自定义用户列表 (替代 QTableWidget)
        self.user_list = UserListWidget()
        
        # 连接用户列表的信号
        self.user_list.edit_user.connect(self.show_edit_dialog)
        self.user_list.device_user.connect(self.show_device_list)
        self.user_list.enter_user.connect(self.enter_user_client)
        self.user_list.toggle_user.connect(self.toggle_user_status)
        self.user_list.delete_user.connect(self.delete_user)
        self.user_list.import_cards_user.connect(self.import_cards_for_user)
        
        card_layout.addWidget(self.user_list, 1)
        
        # 3. 分页
        pagination = QFrame()
        pagination.setFixedHeight(50)
        pagination.setStyleSheet(f"""
            QFrame {{
                background: white;
                border-top: 1px solid {PREMIUM_COLORS['border_light']};
                border-bottom-left-radius: 16px;
                border-bottom-right-radius: 16px;
            }}
        """)
        pagination_layout = QHBoxLayout(pagination)
        pagination_layout.setContentsMargins(16, 0, 16, 0)
        
        self.page_info_label = QLabel()
        self.page_info_label.setStyleSheet(f"color: {PREMIUM_COLORS['text_hint']}; font-size: 12px;")
        pagination_layout.addWidget(self.page_info_label)
        
        pagination_layout.addStretch()
        
        # 翻页按钮
        page_btns = QHBoxLayout()
        page_btns.setSpacing(8)
        
        self.prev_btn = QPushButton("‹")
        self.next_btn = QPushButton("›")
        
        btn_style = f"""
            QPushButton {{
                background: {PREMIUM_COLORS['surface']};
                border: 1px solid {PREMIUM_COLORS['border']};
                border-radius: 14px;
                color: {PREMIUM_COLORS['text_body']};
                font-size: 16px;
                font-weight: bold;
                padding-bottom: 2px;
            }}
            QPushButton:hover {{
                background: {PREMIUM_COLORS['primary']}15;
                color: {PREMIUM_COLORS['primary']};
                border-color: {PREMIUM_COLORS['primary']};
            }}
            QPushButton:disabled {{
                background: {PREMIUM_COLORS['background']};
                color: {PREMIUM_COLORS['border']};
                border-color: {PREMIUM_COLORS['border_light']};
            }}
        """
        
        for btn in [self.prev_btn, self.next_btn]:
            btn.setFixedSize(28, 28)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.setStyleSheet(btn_style)
            btn.clicked.connect(self.change_page)
            
        page_btns.addWidget(self.prev_btn)
        
        self.page_num_label = QLabel("1 / 1")
        self.page_num_label.setStyleSheet(f"""
            color: {PREMIUM_COLORS['text_heading']}; 
            font-weight: 600;
            font-size: 12px;
            padding: 0 8px;
        """)
        page_btns.addWidget(self.page_num_label)
        page_btns.addWidget(self.next_btn)
        
        pagination_layout.addLayout(page_btns)
        card_layout.addWidget(pagination)
        
        layout.addWidget(card, 1)

    def change_page(self):
        sender = self.sender()
        if sender == self.prev_btn:
            self.go_to_page(self.current_page - 1)
        else:
            self.go_to_page(self.current_page + 1)
            
    def on_search(self):
        self.current_page = 1
        self.load_users()
        
    def go_to_page(self, page):
        if 1 <= page <= self.total_pages:
            self.current_page = page
            self.load_users()
            
    def refresh_data(self):
        self.load_users()
        
    def load_users(self):
        keyword = self.search_input.text().strip()
        all_users = self.db_manager.get_all_users(keyword)
        
        # 更新统计
        total_count = len(all_users)
        active_count = sum(1 for u in all_users if u.is_active)
        admin_count = sum(1 for u in all_users if u.role == 'admin')
        
        if "总用户数" in self.stat_cards:
            self.stat_cards["总用户数"].update_value(total_count)
        if "活跃用户" in self.stat_cards:
            self.stat_cards["活跃用户"].update_value(active_count)
        if "管理员" in self.stat_cards:
            self.stat_cards["管理员"].update_value(admin_count)
            
        # 分页处理
        self.total_users = total_count
        self.total_pages = max(1, (self.total_users + self.page_size - 1) // self.page_size)
        
        if self.current_page > self.total_pages:
            self.current_page = self.total_pages
            
        start_idx = (self.current_page - 1) * self.page_size
        end_idx = start_idx + self.page_size
        users = all_users[start_idx:end_idx]
        
        self.update_user_list(users)
        self.update_pagination()
        
    def update_pagination(self):
        start = (self.current_page - 1) * self.page_size + 1
        end = min(self.current_page * self.page_size, self.total_users)
        
        if self.total_users > 0:
            self.page_info_label.setText(f"显示 {start}-{end} 条，共 {self.total_users} 条")
        else:
            self.page_info_label.setText("暂无数据")
            
        self.page_num_label.setText(f"{self.current_page} / {self.total_pages}")
        self.prev_btn.setEnabled(self.current_page > 1)
        self.next_btn.setEnabled(self.current_page < self.total_pages)
        
    def update_user_list(self, users):
        """更新用户列表显示"""
        # 预先计算每个用户的设备数和名片数，避免在组件中逐个查询
        device_counts = {}
        card_counts = {}
        for user in users:
            device_counts[str(user.id)] = Device.objects(user=user).count()
            card_counts[str(user.id)] = Card.objects(user=user).count()
        
        # 调用用户列表组件的方法
        self.user_list.set_users(users, device_counts, card_counts)

    def show_device_list(self, user):
        dialog = DeviceListDialog(self, user)
        dialog.exec()

    def enter_user_client(self, user):
        """进入用户端 - 以该用户身份打开主窗口"""
        from .main_window import MainWindow
        
        # 创建并显示用户主窗口（保留管理后台）
        self.user_main_window = MainWindow(current_user=user)
        self.user_main_window.show()
        self.user_main_window.raise_()
        self.user_main_window.activateWindow()

    def show_add_dialog(self):
        dialog = UserEditDialog(self)
        if dialog.exec():
            self.load_users()
            QMessageBox.information(self, "成功", "用户添加成功")

    def show_edit_dialog(self, user):
        dialog = UserEditDialog(self, user)
        if dialog.exec():
            self.load_users()
            QMessageBox.information(self, "成功", "用户信息已更新")

    def toggle_user_status(self, user):
        new_status = not user.is_active
        action = "启用" if new_status else "禁用"
        
        confirm = QMessageBox.question(
            self, "确认操作",
            f"确定要{action}用户 {user.username} 吗？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if confirm == QMessageBox.StandardButton.Yes:
            if DatabaseManager.update_user(str(user.id), is_active=new_status):
                self.load_users()
                
    def delete_user(self, user):
        if user.username == 'admin':
            QMessageBox.warning(self, "警告", "无法删除超级管理员账号！")
            return
            
        confirm = QMessageBox.warning(
            self, "危险操作",
            f"确定要彻底删除用户 {user.username} 吗？\n\n此操作不可恢复！\n将同时删除该用户的所有：\n1. 名片数据\n2. 链接数据\n3. 填写记录\n4. 设备绑定信息",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if confirm == QMessageBox.StandardButton.Yes:
            if DatabaseManager.delete_user(str(user.id)):
                self.load_users()
                QMessageBox.information(self, "成功", f"用户 {user.username} 已删除")

    def import_cards_for_user(self, user):
        """为指定用户批量导入名片"""
        dialog = BatchImportCardsDialog(user, self)
        if dialog.exec():
            self.load_users()


class BatchImportCardsDialog(QDialog):
    """批量导入名片对话框 - 管理员为用户批量创建名片"""
    
    def __init__(self, target_user, parent=None):
        super().__init__(parent)
        self.target_user = target_user
        self.db_manager = DatabaseManager()
        self.templates = []
        self.categories = []
        self.imported_cards = []
        self._load_templates()
        self._load_categories()
        self.init_ui()
    
    def _load_templates(self):
        """加载所有启用的固定模板"""
        self.templates = self.db_manager.get_all_fixed_templates(is_active=True)
    
    def _load_categories(self):
        """加载目标用户的分类列表"""
        cats = self.db_manager.get_user_categories(self.target_user)
        self.categories = [c.name for c in cats] if cats else []
        if '默认分类' not in self.categories:
            self.categories.insert(0, '默认分类')
    
    def init_ui(self):
        self.setWindowTitle(f"批量导入名片 - {self.target_user.username}")
        self.setMinimumSize(820, 620)
        self.resize(900, 680)
        self.setStyleSheet("QDialog { background-color: white; }")
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        
        # === 1. 顶部 Header ===
        header = QFrame()
        header.setFixedHeight(100)
        header.setStyleSheet(f"""
            QFrame {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 {PREMIUM_COLORS['gradient_green_start']}, 
                    stop:1 {PREMIUM_COLORS['gradient_green_end']});
            }}
        """)
        
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(32, 0, 32, 0)
        header_layout.setSpacing(20)
        
        title_info = QVBoxLayout()
        title_info.setSpacing(6)
        title_info.setAlignment(Qt.AlignmentFlag.AlignVCenter)
        
        title_lbl = QLabel("批量导入名片")
        title_lbl.setStyleSheet("color: white; font-size: 24px; font-weight: 800;")
        
        subtitle_lbl = QLabel(f"为用户 {self.target_user.username} 批量创建名片，支持 CSV 模板导入")
        subtitle_lbl.setStyleSheet("color: rgba(255,255,255,0.85); font-size: 13px; font-weight: 500;")
        
        title_info.addWidget(title_lbl)
        title_info.addWidget(subtitle_lbl)
        
        icon_bg = QLabel("📥")
        icon_bg.setFixedSize(56, 56)
        icon_bg.setAlignment(Qt.AlignmentFlag.AlignCenter)
        icon_bg.setStyleSheet("background: rgba(255,255,255,0.2); border-radius: 28px; font-size: 26px;")
        
        header_layout.addLayout(title_info)
        header_layout.addStretch()
        header_layout.addWidget(icon_bg)
        
        layout.addWidget(header)
        
        # === 2. 操作区域 ===
        body = QWidget()
        body_layout = QVBoxLayout(body)
        body_layout.setContentsMargins(32, 28, 32, 20)
        body_layout.setSpacing(20)
        
        # 步骤说明卡片
        steps_frame = QFrame()
        steps_frame.setStyleSheet(f"""
            QFrame {{
                background: #f0fdf4;
                border: 1px solid #bbf7d0;
                border-radius: 12px;
            }}
        """)
        steps_layout = QVBoxLayout(steps_frame)
        steps_layout.setContentsMargins(20, 16, 20, 16)
        steps_layout.setSpacing(8)
        
        steps_title = QLabel("📋 操作步骤")
        steps_title.setStyleSheet("font-size: 14px; font-weight: 700; color: #166534;")
        steps_layout.addWidget(steps_title)
        
        steps_desc = QLabel(
            "① 点击「下载模板」获取 CSV 模板文件（已包含固定模板字段）\n"
            "② 用 Excel / WPS 打开模板，每行填写一张名片的数据\n"
            "③ 点击「选择文件导入」将填好的 CSV 导入系统"
        )
        steps_desc.setStyleSheet("font-size: 13px; color: #15803d; line-height: 1.6;")
        steps_desc.setWordWrap(True)
        steps_layout.addWidget(steps_desc)
        
        body_layout.addWidget(steps_frame)
        
        # 操作按钮行
        btn_row = QHBoxLayout()
        btn_row.setSpacing(16)
        
        download_btn = GradientButton(
            "⬇ 下载模板",
            PREMIUM_COLORS['gradient_green_start'],
            PREMIUM_COLORS['gradient_green_end']
        )
        download_btn.setFixedSize(160, 44)
        download_btn.clicked.connect(self.download_template)
        btn_row.addWidget(download_btn)
        
        import_btn = GradientButton(
            "📂 选择文件导入",
            PREMIUM_COLORS['gradient_blue_start'],
            PREMIUM_COLORS['gradient_blue_end']
        )
        import_btn.setFixedSize(160, 44)
        import_btn.clicked.connect(self.import_from_file)
        btn_row.addWidget(import_btn)
        
        btn_row.addStretch()
        
        # 模板字段预览按钮
        preview_btn = QPushButton("查看模板字段")
        preview_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        preview_btn.setFixedHeight(36)
        preview_btn.setStyleSheet(f"""
            QPushButton {{
                background: transparent;
                color: {PREMIUM_COLORS['gradient_blue_start']};
                border: 1px solid {PREMIUM_COLORS['gradient_blue_start']}40;
                border-radius: 18px;
                font-size: 13px;
                font-weight: 600;
                padding: 0 16px;
            }}
            QPushButton:hover {{
                background: {PREMIUM_COLORS['gradient_blue_start']}10;
                border-color: {PREMIUM_COLORS['gradient_blue_start']};
            }}
        """)
        preview_btn.clicked.connect(self.show_template_preview)
        btn_row.addWidget(preview_btn)
        
        body_layout.addLayout(btn_row)
        
        # 导入结果区域
        self.result_frame = QFrame()
        self.result_frame.setStyleSheet(f"""
            QFrame {{
                background: #f8fafc;
                border: 1px solid {PREMIUM_COLORS['border_light']};
                border-radius: 12px;
            }}
        """)
        result_layout = QVBoxLayout(self.result_frame)
        result_layout.setContentsMargins(20, 16, 20, 16)
        result_layout.setSpacing(8)
        
        result_header = QLabel("📊 导入结果")
        result_header.setStyleSheet(f"font-size: 14px; font-weight: 700; color: {PREMIUM_COLORS['text_heading']};")
        result_layout.addWidget(result_header)
        
        self.result_label = QLabel("尚未导入数据")
        self.result_label.setStyleSheet(f"font-size: 13px; color: {PREMIUM_COLORS['text_hint']};")
        self.result_label.setWordWrap(True)
        result_layout.addWidget(self.result_label)
        
        self.result_detail = QLabel("")
        self.result_detail.setStyleSheet(f"font-size: 12px; color: {PREMIUM_COLORS['text_body']}; line-height: 1.5;")
        self.result_detail.setWordWrap(True)
        self.result_detail.hide()
        result_layout.addWidget(self.result_detail)
        
        body_layout.addWidget(self.result_frame)
        
        body_layout.addStretch()
        
        layout.addWidget(body, 1)
        
        # === 3. 底部按钮 ===
        bottom = QWidget()
        bottom.setStyleSheet(f"border-top: 1px solid {PREMIUM_COLORS['border_light']};")
        bottom_layout = QHBoxLayout(bottom)
        bottom_layout.setContentsMargins(32, 16, 32, 16)
        
        bottom_layout.addStretch()
        
        close_btn = QPushButton("关闭")
        close_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        close_btn.setFixedSize(100, 40)
        close_btn.setStyleSheet(f"""
            QPushButton {{
                background: white;
                color: {PREMIUM_COLORS['text_body']};
                border: 1px solid {PREMIUM_COLORS['border']};
                border-radius: 20px;
                font-weight: 600;
                font-size: 14px;
            }}
            QPushButton:hover {{
                background: {PREMIUM_COLORS['background']};
                border-color: {PREMIUM_COLORS['text_body']};
            }}
        """)
        close_btn.clicked.connect(self.accept)
        bottom_layout.addWidget(close_btn)
        
        layout.addWidget(bottom)
    
    def _build_csv_header_and_meta(self):
        """构建表头和模板元数据
        
        表头只取 field_name 中第一个别名（'、'分隔的第一个）作为显示名
        
        Returns:
            (header_row, field_meta)
        """
        header = ['名片名称', '分类']
        field_meta = []
        
        for t in self.templates:
            tid = str(t.id)
            full_key = t.field_name
            display_name = full_key.split('、')[0].strip()
            vc = t.value_count or 1
            
            # 解析多值的逐项 placeholder
            per_value_phs = []
            vpt = getattr(t, 'value_placeholder_template', None)
            if vpt and vc > 1:
                import json
                try:
                    parsed = json.loads(vpt)
                    if isinstance(parsed, list):
                        per_value_phs = parsed
                except (json.JSONDecodeError, TypeError):
                    if '{index}' in vpt:
                        per_value_phs = [vpt.replace('{index}', str(i+1)) for i in range(vc)]
            
            default_ph = t.placeholder or ''
            
            if vc <= 1:
                col_name = f"{display_name}\n[{tid}]"
                header.append(col_name)
                field_meta.append({
                    'template_id': tid,
                    'key': full_key,
                    'value_count': 1,
                    'placeholder': default_ph
                })
            else:
                for vi in range(vc):
                    col_name = f"{display_name}__值{vi+1}\n[{tid}]"
                    header.append(col_name)
                    ph = per_value_phs[vi] if vi < len(per_value_phs) and per_value_phs[vi] else default_ph
                    field_meta.append({
                        'template_id': tid,
                        'key': full_key,
                        'value_count': vc,
                        'value_index': vi,
                        'placeholder': ph
                    })
        
        return header, field_meta
    
    def download_template(self):
        """下载 Excel 模板文件（.xlsx），带列宽和分类下拉"""
        from PyQt6.QtWidgets import QFileDialog
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter
        
        if not self.templates:
            QMessageBox.warning(self, "提示", "当前没有启用的固定模板，无法生成模板文件")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "保存名片导入模板",
            f"名片导入模板_{self.target_user.username}.xlsx",
            "Excel 文件 (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            header, field_meta = self._build_csv_header_and_meta()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "名片导入"
            
            # --- 样式定义 ---
            header_font = Font(name='微软雅黑', bold=True, size=10, color="FFFFFF")
            header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
            fixed_col_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
            fixed_col_font = Font(name='微软雅黑', bold=True, size=11, color="1E40AF")
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            border_color = 'C7D2FE'
            thin_border = Border(
                left=Side(style='thin', color=border_color),
                right=Side(style='thin', color=border_color),
                top=Side(style='thin', color=border_color),
                bottom=Side(style='thin', color=border_color)
            )
            
            example_font = Font(name='微软雅黑', color="94A3B8", italic=True, size=10)
            data_font = Font(name='微软雅黑', size=11)
            data_align = Alignment(horizontal="left", vertical="center")
            
            even_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
            
            # --- 写表头（纯文本，名称 + ID 换行，兼容 WPS）---
            for col_idx, col_name in enumerate(header, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                if col_idx <= 2:
                    cell.font = fixed_col_font
                    cell.fill = fixed_col_fill
                else:
                    cell.font = header_font
                    cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border
            
            ws.row_dimensions[1].height = 48
            
            # --- 列宽 ---
            ws.column_dimensions['A'].width = 26
            ws.column_dimensions['B'].width = 16
            for col_idx in range(3, len(header) + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 22
            
            # --- 空白数据行（直接从第2行开始填写）---
            total_rows = 50
            for row_idx in range(2, total_rows + 2):
                for col_idx in range(1, len(header) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value='')
                    cell.border = thin_border
                    cell.font = data_font
                    cell.alignment = data_align
                    if row_idx % 2 == 1:
                        cell.fill = even_fill
                ws.row_dimensions[row_idx].height = 24
            
            # --- 列填写提示（点击单元格时弹出）---
            col_hints = [
                (1, '名片名称', '填写名片名称，每行一张名片'),
            ]
            for col_idx, meta in enumerate(field_meta, 3):
                ph = meta.get('placeholder', '')
                if ph:
                    display = header[col_idx - 1].split('\n')[0] if '\n' in header[col_idx - 1] else header[col_idx - 1]
                    col_hints.append((col_idx, display, ph))
            
            for col_idx, title, hint in col_hints:
                dv = DataValidation(type="custom", formula1="TRUE", allow_blank=True)
                dv.promptTitle = title
                dv.prompt = hint
                dv.showInputMessage = True
                dv.showErrorMessage = False
                col_letter = get_column_letter(col_idx)
                ws.add_data_validation(dv)
                dv.add(f'{col_letter}2:{col_letter}200')
            
            # --- 分类下拉 ---
            if self.categories:
                cat_list = ','.join(self.categories)
                dv = DataValidation(
                    type="list",
                    formula1=f'"{cat_list}"',
                    allow_blank=True,
                    showDropDown=False
                )
                dv.prompt = "请选择分类"
                dv.promptTitle = "分类"
                dv.showInputMessage = True
                dv.showErrorMessage = False
                ws.add_data_validation(dv)
                dv.add('B2:B200')
            
            ws.freeze_panes = 'C2'
            ws.auto_filter.ref = f'A1:{get_column_letter(len(header))}1'
            
            wb.save(file_path)
            
            QMessageBox.information(
                self, "下载成功",
                f"模板已保存至：\n{file_path}\n\n"
                f"共 {len(self.templates)} 个字段\n"
                f"「分类」列可下拉选择\n"
                f"请用 Excel / WPS 打开填写，每行一张名片"
            )
        except Exception as e:
            QMessageBox.critical(self, "错误", f"保存模板失败：{str(e)}")
    
    def import_from_file(self):
        """从 Excel / CSV 文件导入名片"""
        from PyQt6.QtWidgets import QFileDialog
        
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "选择名片文件",
            "",
            "Excel 文件 (*.xlsx);;CSV 文件 (*.csv);;所有文件 (*)"
        )
        
        if not file_path:
            return
        
        try:
            if file_path.lower().endswith('.xlsx'):
                cards_data = self._parse_xlsx(file_path)
            else:
                cards_data = self._parse_csv(file_path)
            
            if not cards_data:
                QMessageBox.warning(self, "提示", "文件中没有有效的名片数据")
                return
            
            preview = ImportPreviewDialog(cards_data, self.templates, self.target_user, self)
            if preview.exec():
                final_data = preview.get_final_data()
                if final_data:
                    result = self.db_manager.batch_create_cards(final_data, self.target_user)
                    self._show_import_result(result)
            
        except UnicodeDecodeError:
            QMessageBox.critical(self, "编码错误", "文件编码不支持，请确保 CSV 以 UTF-8 编码保存")
        except Exception as e:
            QMessageBox.critical(self, "导入失败", f"解析文件时发生错误：{str(e)}")
    
    def _parse_xlsx(self, file_path):
        """解析 Excel (.xlsx) 文件为名片数据列表"""
        from openpyxl import load_workbook
        import json
        
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        
        rows = []
        for row in ws.iter_rows():
            row_data = []
            for cell in row:
                val = cell.value
                if val is not None:
                    # 处理百分比格式，避免如 60% 变成 0.6 的情况
                    if isinstance(val, (int, float)) and cell.number_format and '%' in cell.number_format:
                        val_pct = round(val * 100, 10)
                        if val_pct == int(val_pct):
                            val = f"{int(val_pct)}%"
                        else:
                            val = f"{val_pct}%"
                    row_data.append(str(val))
                else:
                    row_data.append('')
            rows.append(row_data)
        wb.close()
        
        if len(rows) < 2:
            return []
        
        return self._parse_rows(rows)
    
    def _parse_csv(self, file_path):
        """解析 CSV 文件为名片数据列表"""
        import csv, io
        
        content = None
        for encoding in ['utf-8-sig', 'utf-8', 'gbk', 'gb2312', 'gb18030']:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    content = f.read()
                break
            except UnicodeDecodeError:
                continue
        
        if content is None:
            raise UnicodeDecodeError('', b'', 0, 0, '无法识别文件编码')
        
        rows = list(csv.reader(io.StringIO(content)))
        if len(rows) < 2:
            return []
        
        return self._parse_rows(rows)
    
    def _parse_rows(self, rows):
        """通用行数据解析：根据表头中的 [template_id] 匹配模板，提取名片数据"""
        header = rows[0]
        if len(header) < 2:
            return []
        
        # 构建 template_id -> 完整 field_name 的映射
        tpl_map = {str(t.id): t.field_name for t in self.templates}
        
        field_columns = []
        for col_idx in range(2, len(header)):
            col_name = str(header[col_idx]).strip()
            if not col_name:
                continue
            
            template_id = None
            value_index = 0
            
            # 从末尾提取 [template_id]
            if col_name.endswith(']') and '[' in col_name:
                bracket_start = col_name.rindex('[')
                template_id = col_name[bracket_start + 1:-1].strip()
                display_part = col_name[:bracket_start].strip()
            else:
                display_part = col_name
            
            # 解析多值后缀 __值N
            if '__值' in display_part:
                parts = display_part.rsplit('__值', 1)
                try:
                    value_index = int(parts[1]) - 1
                except (ValueError, IndexError):
                    value_index = 0
            
            # 通过 id 查找完整 key，找不到则跳过该列
            if template_id and template_id in tpl_map:
                key = tpl_map[template_id]
            else:
                continue
            
            field_columns.append({
                'col_index': col_idx,
                'template_id': template_id,
                'key': key,
                'value_index': value_index
            })
        
        cards_data = []
        for row_idx in range(1, len(rows)):
            row = rows[row_idx]
            if not row or len(row) < 1:
                continue
            
            name = str(row[0]).strip() if len(row) > 0 else ''
            if not name or name.startswith('(示例)') or name.startswith('示例'):
                continue
            
            category = str(row[1]).strip() if len(row) > 1 else '默认分类'
            
            # 按 template_id 分组（同一个模板的多值列归为一组）
            field_map = {}
            for fc in field_columns:
                col_idx = fc['col_index']
                value = str(row[col_idx]).strip() if col_idx < len(row) else ''
                
                tid = fc['template_id']
                if tid not in field_map:
                    field_map[tid] = {
                        'key': fc['key'],
                        'fixed_template_id': tid,
                        'values': [],
                        'value_count': 1
                    }
                
                fm = field_map[tid]
                vi = fc['value_index']
                while len(fm['values']) <= vi:
                    fm['values'].append('')
                fm['values'][vi] = value
                fm['value_count'] = max(fm['value_count'], vi + 1)
            
            configs = []
            for tid, fm in field_map.items():
                vc = fm['value_count']
                val = fm['values'][0] if (vc <= 1 and fm['values']) else (fm['values'] if vc > 1 else '')
                configs.append({
                    'key': fm['key'],
                    'value': val,
                    'fixed_template_id': fm['fixed_template_id'],
                    'value_count': vc
                })
            
            cards_data.append({
                'name': name,
                'category': category,
                'configs': configs
            })
        
        return cards_data
    
    def _show_import_result(self, result):
        """显示导入结果"""
        success = result['success_count']
        errors = result['error_count']
        error_msgs = result.get('errors', [])
        
        if errors == 0:
            self.result_label.setText(f"✅ 导入完成！成功创建 {success} 张名片")
            self.result_label.setStyleSheet("font-size: 14px; color: #16a34a; font-weight: 600;")
        else:
            self.result_label.setText(f"⚠️ 导入完成：成功 {success} 张，失败 {errors} 张")
            self.result_label.setStyleSheet(f"font-size: 14px; color: {PREMIUM_COLORS['gradient_gold_start']}; font-weight: 600;")
        
        if error_msgs:
            self.result_detail.setText('\n'.join(error_msgs[:10]))
            if len(error_msgs) > 10:
                self.result_detail.setText(self.result_detail.text() + f"\n... 还有 {len(error_msgs) - 10} 条错误")
            self.result_detail.show()
        else:
            self.result_detail.hide()
    
    def show_template_preview(self):
        """预览模板字段"""
        if not self.templates:
            QMessageBox.information(self, "模板字段", "当前没有启用的固定模板")
            return
        
        dialog = QDialog(self)
        dialog.setWindowTitle("模板字段预览")
        dialog.setMinimumSize(600, 450)
        dialog.setStyleSheet("QDialog { background: white; }")
        
        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(24, 24, 24, 24)
        layout.setSpacing(16)
        
        title = QLabel(f"固定模板字段（共 {len(self.templates)} 个）")
        title.setStyleSheet(f"font-size: 16px; font-weight: 700; color: {PREMIUM_COLORS['text_heading']};")
        layout.addWidget(title)
        
        hint = QLabel("以下字段将出现在 CSV 模板中，表头格式为 [模板ID]字段名")
        hint.setStyleSheet(f"font-size: 12px; color: {PREMIUM_COLORS['text_hint']};")
        layout.addWidget(hint)
        
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setStyleSheet("QScrollArea { background: transparent; border: none; }")
        
        content = QWidget()
        content_layout = QVBoxLayout(content)
        content_layout.setContentsMargins(0, 0, 0, 0)
        content_layout.setSpacing(8)
        
        for i, t in enumerate(self.templates, 1):
            row = QFrame()
            row.setStyleSheet(f"""
                QFrame {{
                    background: #f8fafc;
                    border: 1px solid {PREMIUM_COLORS['border_light']};
                    border-radius: 8px;
                }}
            """)
            row_layout = QHBoxLayout(row)
            row_layout.setContentsMargins(12, 10, 12, 10)
            row_layout.setSpacing(12)
            
            idx_label = QLabel(str(i))
            idx_label.setFixedWidth(24)
            idx_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            idx_label.setStyleSheet(f"""
                background: {PREMIUM_COLORS['gradient_blue_start']}15;
                color: {PREMIUM_COLORS['gradient_blue_start']};
                border-radius: 12px;
                font-size: 11px;
                font-weight: 700;
                padding: 2px;
            """)
            row_layout.addWidget(idx_label)
            
            info_layout = QVBoxLayout()
            info_layout.setSpacing(2)
            
            name_label = QLabel(t.field_name)
            name_label.setStyleSheet(f"font-size: 13px; font-weight: 600; color: {PREMIUM_COLORS['text_heading']};")
            info_layout.addWidget(name_label)
            
            meta_parts = [f"ID: {str(t.id)[:8]}..."]
            if t.value_count and t.value_count > 1:
                meta_parts.append(f"多值×{t.value_count}")
            if t.placeholder:
                meta_parts.append(f"提示: {t.placeholder}")
            if t.category:
                meta_parts.append(f"分类: {t.category}")
            
            meta_label = QLabel(" · ".join(meta_parts))
            meta_label.setStyleSheet(f"font-size: 11px; color: {PREMIUM_COLORS['text_hint']};")
            info_layout.addWidget(meta_label)
            
            row_layout.addLayout(info_layout, 1)
            content_layout.addWidget(row)
        
        content_layout.addStretch()
        scroll.setWidget(content)
        layout.addWidget(scroll, 1)
        
        close_btn = QPushButton("关闭")
        close_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        close_btn.setFixedSize(80, 36)
        close_btn.setStyleSheet(f"""
            QPushButton {{
                background: {PREMIUM_COLORS['background']};
                border: 1px solid {PREMIUM_COLORS['border']};
                border-radius: 18px;
                color: {PREMIUM_COLORS['text_body']};
                font-weight: 600;
            }}
            QPushButton:hover {{ background: {PREMIUM_COLORS['border_light']}; }}
        """)
        close_btn.clicked.connect(dialog.accept)
        
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        btn_layout.addWidget(close_btn)
        layout.addLayout(btn_layout)
        
        dialog.exec()


class ImportPreviewDialog(QDialog):
    """导入预览对话框 - 解析后预览、编辑、确认再导入"""
    
    def __init__(self, cards_data, templates, target_user, parent=None):
        super().__init__(parent)
        self.cards_data = cards_data
        self.templates = templates
        self.target_user = target_user
        self._confirmed = False
        
        # 构建模板 ID -> 显示名映射
        self.tpl_display = {}
        for t in templates:
            self.tpl_display[str(t.id)] = t.field_name.split('、')[0].strip()
        
        # 收集所有出现的字段列（按 template_id 去重，保持顺序）
        self.field_columns = []
        seen_tids = set()
        for card in cards_data:
            for cfg in card.get('configs', []):
                tid = cfg.get('fixed_template_id', '')
                if tid and tid not in seen_tids:
                    seen_tids.add(tid)
                    self.field_columns.append({
                        'template_id': tid,
                        'display': self.tpl_display.get(tid, cfg.get('key', '').split('、')[0]),
                        'value_count': cfg.get('value_count', 1)
                    })
        
        self.init_ui()
        self._populate_table()
    
    def init_ui(self):
        self.setWindowTitle(f"导入预览 - {self.target_user.username}")
        self.setMinimumSize(960, 580)
        self.resize(1100, 680)
        self.setStyleSheet("QDialog { background-color: white; }")
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        
        # --- 顶部 Header ---
        header = QFrame()
        header.setFixedHeight(72)
        header.setStyleSheet(f"""
            QFrame {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 {PREMIUM_COLORS['gradient_blue_start']}, stop:1 {PREMIUM_COLORS['gradient_blue_end']});
            }}
        """)
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(28, 0, 28, 0)
        
        title = QLabel(f"📋 预览导入数据 — 共 {len(self.cards_data)} 张名片")
        title.setStyleSheet("color: white; font-size: 16px; font-weight: 700;")
        
        hint = QLabel("双击单元格可编辑，确认无误后点击「确认导入」")
        hint.setStyleSheet("color: rgba(255,255,255,0.75); font-size: 12px;")
        
        title_box = QVBoxLayout()
        title_box.setSpacing(2)
        title_box.addWidget(title)
        title_box.addWidget(hint)
        header_layout.addLayout(title_box)
        header_layout.addStretch()
        layout.addWidget(header)
        
        # --- 工具栏 ---
        toolbar = QFrame()
        toolbar.setFixedHeight(48)
        toolbar.setStyleSheet(f"background: {PREMIUM_COLORS['background']}; border-bottom: 1px solid {PREMIUM_COLORS['border_light']};")
        tb_layout = QHBoxLayout(toolbar)
        tb_layout.setContentsMargins(16, 0, 16, 0)
        tb_layout.setSpacing(12)
        
        self.count_label = QLabel(f"共 {len(self.cards_data)} 条")
        self.count_label.setStyleSheet(f"color: {PREMIUM_COLORS['text_body']}; font-size: 13px; font-weight: 600;")
        tb_layout.addWidget(self.count_label)
        
        tb_layout.addStretch()
        
        del_btn = QPushButton("🗑 删除选中行")
        del_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        del_btn.setStyleSheet(f"""
            QPushButton {{
                background: white; border: 1px solid {PREMIUM_COLORS['border']};
                border-radius: 6px; padding: 4px 14px; font-size: 12px;
                color: {PREMIUM_COLORS['text_body']};
            }}
            QPushButton:hover {{ border-color: #ef4444; color: #ef4444; }}
        """)
        del_btn.clicked.connect(self._delete_selected)
        tb_layout.addWidget(del_btn)
        
        layout.addWidget(toolbar)
        
        # --- 表格 ---
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setStyleSheet(f"""
            QTableWidget {{
                border: none;
                gridline-color: {PREMIUM_COLORS['border_light']};
                font-size: 12px;
                alternate-background-color: #f8fafc;
            }}
            QTableWidget::item {{
                padding: 6px 10px;
            }}
            QTableWidget::item:selected {{
                background: {PREMIUM_COLORS['gradient_blue_start']}18;
                color: {PREMIUM_COLORS['text_heading']};
            }}
            QHeaderView::section {{
                background: {PREMIUM_COLORS['background']};
                border: none;
                border-bottom: 2px solid {PREMIUM_COLORS['border']};
                padding: 8px 10px;
                font-size: 12px;
                font-weight: 700;
                color: {PREMIUM_COLORS['text_heading']};
            }}
        """)
        layout.addWidget(self.table, 1)
        
        # --- 底部按钮 ---
        bottom = QFrame()
        bottom.setFixedHeight(64)
        bottom.setStyleSheet(f"background: white; border-top: 1px solid {PREMIUM_COLORS['border_light']};")
        btn_layout = QHBoxLayout(bottom)
        btn_layout.setContentsMargins(24, 0, 24, 0)
        btn_layout.setSpacing(12)
        
        btn_layout.addStretch()
        
        cancel_btn = QPushButton("取消")
        cancel_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        cancel_btn.setFixedSize(100, 38)
        cancel_btn.setStyleSheet(f"""
            QPushButton {{
                background: white;
                border: 1px solid {PREMIUM_COLORS['border']};
                border-radius: 8px;
                color: {PREMIUM_COLORS['text_body']};
                font-size: 13px; font-weight: 600;
            }}
            QPushButton:hover {{ background: #f1f5f9; }}
        """)
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)
        
        confirm_btn = QPushButton(f"✅ 确认导入 ({len(self.cards_data)} 张)")
        confirm_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        confirm_btn.setFixedHeight(38)
        confirm_btn.setMinimumWidth(160)
        confirm_btn.setStyleSheet(f"""
            QPushButton {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {PREMIUM_COLORS['gradient_blue_start']}, stop:1 {PREMIUM_COLORS['gradient_blue_end']});
                border: none; border-radius: 8px;
                color: white; font-size: 13px; font-weight: 700;
                padding: 0 24px;
            }}
            QPushButton:hover {{ opacity: 0.9; }}
        """)
        confirm_btn.clicked.connect(self._confirm_import)
        self.confirm_btn = confirm_btn
        btn_layout.addWidget(confirm_btn)
        
        layout.addWidget(bottom)
    
    def _build_col_headers(self):
        """构建表格列头：序号 + 名片名称 + 分类 + 各字段列"""
        headers = ['序号', '名片名称', '分类']
        for fc in self.field_columns:
            vc = fc['value_count']
            if vc <= 1:
                headers.append(fc['display'])
            else:
                for vi in range(vc):
                    headers.append(f"{fc['display']}__值{vi+1}")
        return headers
    
    def _populate_table(self):
        """填充表格数据"""
        headers = self._build_col_headers()
        
        self.table.setColumnCount(len(headers))
        self.table.setRowCount(len(self.cards_data))
        self.table.setHorizontalHeaderLabels(headers)
        
        h = self.table.horizontalHeader()
        h.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(0, 48)
        h.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        h.setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(2, 100)
        for ci in range(3, len(headers)):
            h.setSectionResizeMode(ci, QHeaderView.ResizeMode.Interactive)
            self.table.setColumnWidth(ci, 140)
        
        for row_idx, card in enumerate(self.cards_data):
            # 序号（只读）
            idx_item = QTableWidgetItem(str(row_idx + 1))
            idx_item.setFlags(idx_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            idx_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            idx_item.setForeground(QColor(PREMIUM_COLORS['text_hint']))
            self.table.setItem(row_idx, 0, idx_item)
            
            # 名片名称
            self.table.setItem(row_idx, 1, QTableWidgetItem(card.get('name', '')))
            
            # 分类
            self.table.setItem(row_idx, 2, QTableWidgetItem(card.get('category', '默认分类')))
            
            # 构建 template_id -> config 映射
            cfg_map = {}
            for cfg in card.get('configs', []):
                tid = cfg.get('fixed_template_id', '')
                if tid:
                    cfg_map[tid] = cfg
            
            # 填充字段列
            col_offset = 3
            for fc in self.field_columns:
                tid = fc['template_id']
                vc = fc['value_count']
                cfg = cfg_map.get(tid, {})
                val = cfg.get('value', '')
                
                if vc <= 1:
                    display_val = val if isinstance(val, str) else str(val)
                    self.table.setItem(row_idx, col_offset, QTableWidgetItem(display_val))
                    col_offset += 1
                else:
                    if isinstance(val, list):
                        values = val
                    else:
                        values = [val] if val else []
                    for vi in range(vc):
                        v = values[vi] if vi < len(values) else ''
                        self.table.setItem(row_idx, col_offset, QTableWidgetItem(str(v)))
                        col_offset += 1
    
    def _delete_selected(self):
        """删除选中的行"""
        rows = sorted(set(idx.row() for idx in self.table.selectedIndexes()), reverse=True)
        if not rows:
            return
        
        for r in rows:
            if r < len(self.cards_data):
                self.cards_data.pop(r)
            self.table.removeRow(r)
        
        # 刷新序号
        for i in range(self.table.rowCount()):
            item = self.table.item(i, 0)
            if item:
                item.setText(str(i + 1))
        
        count = self.table.rowCount()
        self.count_label.setText(f"共 {count} 条")
        self.confirm_btn.setText(f"✅ 确认导入 ({count} 张)")
    
    def _confirm_import(self):
        """确认导入"""
        if self.table.rowCount() == 0:
            QMessageBox.warning(self, "提示", "没有可导入的数据")
            return
        self._confirmed = True
        self.accept()
    
    def get_final_data(self):
        """从表格中读取最终数据（含用户编辑）"""
        if not self._confirmed:
            return None
        
        final = []
        for row_idx in range(self.table.rowCount()):
            name = (self.table.item(row_idx, 1).text().strip() 
                    if self.table.item(row_idx, 1) else '')
            if not name:
                continue
            
            category = (self.table.item(row_idx, 2).text().strip()
                       if self.table.item(row_idx, 2) else '默认分类')
            
            configs = []
            col_offset = 3
            for fc in self.field_columns:
                tid = fc['template_id']
                vc = fc['value_count']
                
                # 从原始数据中获取完整 key
                original_key = None
                if row_idx < len(self.cards_data):
                    for cfg in self.cards_data[row_idx].get('configs', []):
                        if cfg.get('fixed_template_id') == tid:
                            original_key = cfg.get('key')
                            break
                if not original_key:
                    original_key = self.tpl_display.get(tid, tid)
                
                if vc <= 1:
                    val = (self.table.item(row_idx, col_offset).text().strip()
                          if self.table.item(row_idx, col_offset) else '')
                    col_offset += 1
                else:
                    val = []
                    for vi in range(vc):
                        v = (self.table.item(row_idx, col_offset).text().strip()
                            if self.table.item(row_idx, col_offset) else '')
                        val.append(v)
                        col_offset += 1
                
                configs.append({
                    'key': original_key,
                    'value': val,
                    'fixed_template_id': tid,
                    'value_count': vc
                })
            
            final.append({
                'name': name,
                'category': category or '默认分类',
                'configs': configs
            })
        
        return final
